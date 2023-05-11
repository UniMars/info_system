# -*- coding: utf-8 -*-
import json
import logging
import os.path
import random
import re
import time
import traceback
import urllib
from datetime import datetime, timedelta
from urllib import request as urllib2
from urllib.parse import quote

import openpyxl as op
import pandas as pd
import xlwt
from bs4 import BeautifulSoup
from django.conf import settings
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.utils.exceptions import IllegalCharacterError
from selenium import webdriver
from selenium.common import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait

logger = logging.getLogger('django')


class TouTiaoContent:
    """
    一个根据指定关键词抓取今日头条（中国新闻平台）文章的类。
    属性:
        root_path (str): 保存输出文件的根路径。
        driver_path (str): selenium webdriver的路径。
        area_keyword (str): 搜索的区域关键词，会以今日头条+area_keyword命名保存xlsx文件。
        *args: 其他搜索关键词。
    方法:
        append_to_worksheet(data): 将提取的文章数据附加到工作表。
        extract_article_info(url): 从文章的URL中提取相关信息。
        toutiao_spider(): 开始网络抓取过程的主要方法。
        parse_time_str(time_str): 将时间字符串转换为pd.Timestamp对象。
        judge_exist(search_content): 检查页面上是否存在元素，如有必要进行刷新。
    """

    def __init__(self, root_path="", driver_path="", area_keyword="", search_args=None):
        self._start_date = pd.to_datetime(datetime.now())  # TODO 读取最新日期
        self._area_keyword = area_keyword
        search_args = search_args if search_args else []
        self._search_keywords = search_args
        self._search_keyword = ' '.join([str(i) for i in search_args])
        if self._area_keyword:
            self._search_keyword = f'{self._area_keyword} {self._search_keyword}'
        file_name = f'今日头条+{area_keyword}.xlsx'
        if not root_path:
            root_path = settings.BASE_DIR / "DATA/头条/"
        if not os.path.exists(root_path):
            os.mkdir(root_path)
        self._file_path = root_path / file_name
        try:
            wb = op.load_workbook(self._file_path)  # 该文件需要提前创建
        except FileNotFoundError:
            if not os.path.exists(root_path):
                os.makedirs(root_path)
            wb = op.Workbook()
            sh = wb.worksheets[0]
            sh.append(['title', 'url', 'area_keyword', 'search_keyword', 'date', 'media', 'context', 'comment'])
            wb.save(self._file_path)
            wb = op.load_workbook(self._file_path)
        self._wb = wb
        self._ws = wb.worksheets[0]
        option = webdriver.ChromeOptions()
        prefs = {"profile.managed_default_content_settings.images": 2}  # 设置无图模式
        option.add_experimental_option("prefs", prefs)  # 加载无图模式设置
        # option.add_argument('--incognito')  # 无痕浏览
        # option.add_argument(
        # 'user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/112.0.0.0 Safari/537.36')  # 浏览器标头
        option.add_experimental_option('excludeSwitches', ['enable-automation'])
        option.add_experimental_option('useAutomationExtension', False)
        if not driver_path:
            driver_path = 'C:/ProgramData/selenium/chromedriver.exe'
        driver = webdriver.Chrome(executable_path=driver_path, options=option)
        driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
            "source": """
            Object.defineProperty(navigator, 'webdriver', {
              get: () => false
            })
          """
        })
        self._driver = driver

    def append_to_worksheet(self, data):
        if data:
            try:
                self._ws.append(data)
            except IllegalCharacterError:
                for _ in range(len(data)):
                    if isinstance(data[_], str):
                        data[_] = ILLEGAL_CHARACTERS_RE.sub('□', data[_])  # 去除非法字符
                self._ws.append(data)

    def extract_article_info(self, url):
        title = ''
        date = ''
        media = ''
        context = ''
        comment = ''
        for _ in range(3):
            try:
                self._driver.get(url)  # 获取具体的文章页面
                # time.sleep(0.25)
                WebDriverWait(self._driver, timeout=10).until(
                    EC.presence_of_element_located((By.CLASS_NAME, 'article-content')))
                tdmc = self._driver.find_element(By.CLASS_NAME, 'article-content').text
                list1 = tdmc.split('\n', 2)
                title = list1[0]
                date = list1[1].split('·')[0]
                media = list1[1].split('·')[1]
                context = list1[2] if len(list1) > 2 else ""
                # 解析网页获取文章的题目、时间、来源、正文、部分评论
                comment = self._driver.find_element(By.CLASS_NAME, 'comment-list').text
            except TimeoutException:
                continue
            except Exception as _:
                logger.warning(f"Error:{_}")
                time.sleep(3000)
                continue
                # break
        if title == '' and date == '':
            return None
        return [title, url, self._area_keyword, self._search_keyword, date, media, context, comment]

    def toutiao_spider(self):
        start_time = time.time()
        url_list = []  # TODO
        for j in range(18):  # 今日头条搜索资讯一栏大概每个关键词会弹出18页
            try:  # 为了防止某搜索页面没有18页或者第18页没有10个资讯而报错
                page_url = 'https://so.toutiao.com/search?keyword=' \
                           + quote(self._search_keyword) \
                           + '&pd=information&source=search_subtab_switch&dvpf=pc&aid=4916&page_num=' + str(j)
                # url链接
                self._driver.get(page_url)
                time.sleep(3)  # 可根据实际情况调整
                target_position = '//div[@data-i]/div/div/div/div/div[1]/div/a'
                target = self.judge_exist({'by': By.XPATH, 'value': target_position})
                if not target:
                    break

                for item in self._driver.find_elements(By.XPATH, '//div[@data-i]'):
                    for _ in item.find_elements(By.XPATH, 'div/div/div/div/div[last()]/div[last()]/div[last()]'):
                        # print(timetext.text)
                        time_str = _.text.split('\n')[-1]
                        date = self.parse_time_str(time_str)
                        if date < self._start_date:
                            print("日期超出范围")  # TODO 日期超出范围
                    for link in item.find_elements(By.XPATH, 'div/div/div/div/div[1]/div/a'):
                        url_list.append(link.get_attribute('href'))
            except Exception as _:
                logger.error(_)
        url_count = 0
        for url in url_list:
            if url_count % 10 == 0:
                page = url_count // 10
                logger.debug(f"{datetime.now()}:已爬取到关键词:{self._search_keyword}的第{page}页")  # 查看爬取进度
            url_count += 1
            data = self.extract_article_info(url)
            self.append_to_worksheet(data)
            # self._driver.back()  # 退回搜索文章的页面
            time.sleep(0.15)
        self._driver.close()
        self._wb.save(self._file_path)
        end_time = time.time()
        minute = int((end_time - start_time) // 60)
        second = int((end_time - start_time) % 60)
        logger.debug(f"{self._area_keyword} {self._search_keyword} 爬取完成：共耗时{minute}min{second}s")

    @staticmethod
    def parse_time_str(self, time_str):
        try:
            print(time_str)
            if '分钟前' in time_str:
                pub_date = datetime.now() - timedelta(minutes=int(time_str.split('分钟前')[0]))
                pub_date = pd.Timestamp(pub_date.replace(second=0, microsecond=0))
            elif '小时前' in time_str:
                pub_date = datetime.now() - timedelta(hours=int(time_str.split('小时前')[0]))
                pub_date = pd.Timestamp(pub_date.replace(second=0, microsecond=0))
            elif '天前' in time_str:
                pub_date = datetime.now() - timedelta(days=int(time_str.split('天前')[0]))
                pub_date = pd.Timestamp(pub_date.replace(hour=0, minute=0, second=0, microsecond=0))
            elif '昨天' in time_str:
                pub_date = pd.to_datetime(time_str.split(' ')[-1], errors='coerce') - timedelta(days=1)
            elif '前天' in time_str:
                pub_date = pd.to_datetime(time_str.split(' ')[-1], errors='coerce') - timedelta(days=2)
            else:
                pub_date = pd.to_datetime(time_str, format="%m月%d日", errors='coerce')
                pub_date = pub_date.replace(
                    year=datetime.now().year) if pub_date is not pd.NaT else pub_date
                pub_date = pd.to_datetime(time_str, format="%Y年%m月%d日",
                                          errors='coerce') if pub_date is pd.NaT else pub_date
            if pub_date is pd.NaT:
                pub_date = None
            print(pub_date)
        except Exception as _:
            pub_date = None
            print(_)
        return pub_date

    def judge_exist(self, search_content):
        for _ in range(10):
            target = self._driver.find_elements(**search_content)
            if target:
                return target
            time.sleep(10)
            self._driver.refresh()
        target = self._driver.find_elements(**search_content)
        return target


class WeiboContent:
    """
    @author: 李梓琪
    注意事项：
    1、更换cookie和agents序列
    2、程序入口处可选择需要的关键词列表和省份
    3、默认从话题第一页开始爬取，有需要可以更改页数
    4、最终print的dic1，会显示各个关键词爬取的最终页数。部分爬取过程中超时报错的，可以从显示的最终页数开始重新爬取。
    5、其他见代码注释。
    """

    # 爬取代码
    @staticmethod
    def read_html(key, page):  # 关键词，从话题第几页开始爬取
        info_list = []
        try:
            info_list = []  # 存储所有微博的数据

            my_agents = ["", "", "", "", ""]  # 设置一系列代理，通过random函数更换代理，减少反爬屏蔽
            randdom_agents = random.choice(my_agents)
            cookies = ''
            headers = {'User-Agent': randdom_agents, 'cookie': cookies}

            key_word = urllib.parse.quote(key)  # 将关键词转化为url格式

            while page:  # 爬取每一页的数据
                flag = 0  # 等于1时跳出循环

                url = 'https://m.weibo.cn/api/container/getIndex?containerid=100103type%3D61%26q%3D{0}%26t%3D0&page_type=searchall'.format(
                    key_word)
                if page > 1:  # 第一页和其他话题页url有些许区别。爬取的是网页端实时微博下拉加载页，数据格式为键值对。可根据需要爬取的页面替换url。
                    url += '&page={0}'.format(page)

                request = urllib2.Request(url=url, headers=headers)
                req = urllib2.urlopen(request, timeout=10)  # 发送请求
                index_doc = req.read()  # 读取页面内容
                index_json = json.loads(index_doc)  # 使用json.loads()将获取的返回内容转换成json格式

                cards = index_json['data']['cards']  # 微博数据集。如需爬取综合排序、热门微博等数据，需要根据具体数据层级来获得微博数据集。
                start = 0
                end = len(cards)  # 该页面微博数量
                if end == 0:  # 该页面无数据，数据爬取结束，跳出循环
                    print('no more data')
                    break
                for i in range(start, end):
                    mblog_url = cards[i]['scheme']  # 该页面第i-1条微博的地址
                    mblog = cards[i]['mblog']  # 该页面第i-1条微博

                    if 'retweeted_status' in mblog:  # 跳过转发的微博。需要保留可删除本条。
                        print('no content {0}'.format(i))
                        continue

                    reposts_count = mblog['reposts_count']  # 转发数量
                    comments_count = mblog['comments_count']  # 评论数量
                    attitudes_count = mblog['attitudes_count']  # 点赞数量

                    date = mblog['created_at']  # 发表时间
                    month = date[4:7]  # 从时间中提取月份
                    year = date[-4:]  # 从时间中提取年份
                    if '2019' in date or '2018' in date:  # 只爬取2020及以后的数据，所以出现19、18年的数据就跳出循环，停止爬取数据
                        flag = 1
                        break

                    mid = mblog['id']  # 微博ID
                    pic_num = mblog['pic_num']  # 微博带图数量

                    long_text = mblog['isLongText']  # 判断是否为长微博，如果是的话需要进行处理
                    if long_text:
                        text_href = 'https://m.weibo.cn/statuses/extend?id={0}'.format(mblog['id'])  # 长微博地址
                        text_req = urllib2.Request(url=text_href, headers=headers)
                        text_doc = urllib2.urlopen(text_req, timeout=10).read()  # 读取长微博数据
                        text_json = json.loads(text_doc)  # 转换为json格式
                        text = text_json['data']['longTextContent']  # 长微博文本数据
                    else:
                        text = mblog['text']  # 微博文本数据
                    length = mblog['textLength']  # 微博文本长度
                    text_soup = BeautifulSoup(text, "html.parser")  # 对文本数据进行处理，转化为txt格式
                    content = text_soup.get_text()  # txt文本
                    urls = len(re.findall('small_web_default.png', text)) + len(
                        re.findall('small_article_default.png', text))  # 文本中超链接数量

                    source = mblog['source']  # 微博数据来源

                    is_location = len(re.findall('small_location_default.png', text))  # 微博是否带有定位信息
                    location = text_soup.find_all("span", {"class": "surl-text"})[
                        -1].get_text() if is_location else ""  # 定位

                    # 提取用户信息
                    user = mblog['user']  # 用户数据
                    user_id = user['id']
                    user_name = user['screen_name']
                    user_gender = user['gender']
                    user_statuses_count = user['statuses_count']  # 发博总数
                    user_followers_count = user['followers_count']  # 粉丝数
                    user_follow_count = user['follow_count']  # 关注数
                    user_verified = user['verified']  # 微博是否认证
                    user_verified_type = user['verified_type']  # 微博认证类型，-1普通用户 0名人 1政府 2企业 3媒体 4校园
                    user_verified_reason = user['verified_reason'] if 'verified_reason' in user else ""  # 认证原因
                    user_urank = user['urank']  # 微博等级

                    is_splmt = 0
                    splmt_type = ""
                    splmt_title = ""
                    splmt_url = ""
                    if 'page_info' in mblog:  # 是否链接其他页面
                        page_info = mblog['page_info']
                        is_splmt = 1
                        splmt_type = page_info['type']  # 视频、直播、网页、股票等
                        splmt_title = page_info["page_title"]
                        splmt_url = page_info["page_url"]

                    # 单条微博的信息
                    info = (year, month, mblog_url, mid, user_id, user_name, source, user_gender, user_statuses_count,
                            user_followers_count, user_follow_count,
                            user_verified, user_verified_type, user_verified_reason, user_urank, location, content,
                            length, is_splmt, splmt_type, splmt_title, splmt_url, urls,
                            reposts_count, comments_count, attitudes_count, pic_num)
                    info_list.append(info)
                    print('card {0} fin'.format(i))
                print('get page:%d' % page)

                page += 1  # 爬取下一页数据
                time.sleep(2 * random.random())  # 停两秒，反爬
                if flag == 1:
                    print('datas needed are completed')
                    break
            return info_list, page
        except Exception as _:
            print(_)
            traceback.print_exc()
            return info_list, page

    # 存入excel
    def save_excel(self, key="", page=1):  # 关键词，从话题第几页开始爬取
        info_list, page2 = self.read_html(key, page)  # 返回爬取的数据和页面
        saved_file = xlwt.Workbook()
        sh = saved_file.add_sheet(u'Sheet', cell_overwrite_ok=True)
        sh.col(2).width = 50 * 80
        sh.col(21).width = 50 * 80
        sh.col(16).width = 256 * 80
        style = xlwt.easyxf('align: wrap on, vert centre, horiz center')  # 设置显示格式

        row0 = (
            'time', 'month', 'mblogurl', 'mid', 'user_id', 'user_name', 'source', 'user_gender', 'user_statuses_count',
            'user_followers_count', 'user_follow_count', 'user_verified', 'user_verified_type', 'user_verified_reason',
            'user_urank',
            'location', 'content', 'length', 'is_splmt', 'splmt_type', 'splmt_title', 'splmt_url', 'urls_num',
            'reposts_count', 'comments_count', 'attitudes_count', 'pic_num')
        for i in range(0, len(row0)):  # 写入第一行标题
            sh.write(0, i, row0[i], style)

        for i in range(0, len(info_list)):  # 将info_list中的微博逐行写入
            for j in range(0, len(info_list[i])):
                sh.write(i + 1, j, info_list[i][j], style)

        saved_file.save('D:/北京/{0}.xls'.format(key))  # 保存文件
        print('save success')
        return page2
